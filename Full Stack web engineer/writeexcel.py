from openpyxl import *
from datetime import date, timedelta

def excel(l,date) :
      
      wb = load_workbook('Attendance.xlsx')
      ws = wb.worksheets[0]
      
      
      ws.cell(2,5,'Pointeuse '+str(date))
      row=5
      for i in l :
         ws.cell(row,1,str(i[1])) 
         ws.cell(row,2,str(i[2]))
         ws.cell(row,3,str(i[3])[:11])
         ws.cell(row,4,str(i[3])[11:])
         ws.cell(row,5,str(i[4])[11:])
         ws.cell(row,6,str(i[5]+1))
         ws.cell(row,7,str(i[5]))
         if i[5]-8<0 :
             ws.cell(row,8,str(i[5]-8))
         else :
             ws.cell(row,9,str(i[5]-8))
         
         if str(i[6]) == 'None' :
            ws.cell(row,10,'')
            ws.cell(row,11,'')
         else:
             ws.cell(row,10,str(i[6]))
             ws.cell(row,11,str(i[6] - i[5]))
             
         if str(i[7]) == 'None' :
            ws.cell(row,12,'')
         else:
             ws.cell(row,12,str(i[7]))

        
         row+=1
      wb.save('Attendance1.xlsx') 
def excelmonth(l) :
      wb = load_workbook('monthly_recap.xlsx')
      ws = wb.worksheets[1]
      row =2
      print l
      for i in l :
         k=round(i[2],2)
         if (int(str(k)[(str(k).find('.'))+1])>=6) :
           k=k-0.6+1
           ws.cell(row,6,str(k))
         else :
            ws.cell(row,6,str(k))
  	     
         ws.cell(row,1,i[0])
         ws.cell(row,2,i[1])
         row+=1
      wb.save('monthly_recap1.xlsx')
#          ws.cell(row,1,i[0])
#          ws.cell(row,2,i[1])
#	  k=round(i[2],2)
#          ws.cell(row,6,str(k))
#          row+=1
       
